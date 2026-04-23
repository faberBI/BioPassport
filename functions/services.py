# ======================================================
# ======================================================
import os
import json
import base64
import hashlib
import unicodedata

from io import BytesIO
from datetime import datetime, timezone
from contextlib import contextmanager
from urllib.parse import urlparse
from difflib import get_close_matches
from typing import Optional, Dict

import streamlit as st
import pandas as pd
import pdfplumber
from PIL import Image
from openai import OpenAI
import qrcode
import requests

from cryptography.fernet import Fernet
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import load_workbook

# ======================================================
# CONFIG
# ======================================================

PASSPORT_DIR = "passports"
EXCEL_FILE = os.path.join(PASSPORT_DIR, "passport_archive.xlsx")  # fallback legacy (non usato se DB attivo)

PRODUCT_FIELDS = {
    "mobile": {
        "pdf": [
            "Nome prodotto",
            "Numero di modello",
            "Produttore",
            "Materiali/componenti utilizzati",
            "Percentuale di contenuto riciclato",
            "Sostanze preoccupanti",
            "Energia consumata",
            "Luogo di Produzione",
            "Durabilità",
            "Istruzioni di riparazione",
            "Parti sostituibili",
            "Indicazioni di smaltimento",
            "Fine vita",
            "Prezzo in euro",
            "Peso",
            "Dimensioni",
            "GTIN",
        ],
        "image": ["Colore", "Condizioni"],
    },
    "lampada": {"pdf": [], "image": []},
    "bicicletta": {"pdf": [], "image": []},
}


# ======================================================
# UTILS
# ======================================================

def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize(text: str) -> str:
    """
    ✅ versione richiesta: NFKD / ascii / underscore
    """
    import unicodedata
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", str(text))
    text = text.encode("ascii", "ignore").decode()
    return text.lower().strip().replace(" ", "_")


def compute_sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _canonical_json(obj) -> bytes:
    return json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def compute_passport_hash(passport: dict) -> str:
    tmp = dict(passport)
    tmp.pop("digital_signature", None)
    return hashlib.sha256(_canonical_json(tmp)).hexdigest()


def get_default_issuer() -> dict:
    return {
        "legal_name": os.getenv("ISSUER_LEGAL_NAME", "Nuvia S.r.l."),
        "country": os.getenv("ISSUER_COUNTRY", "IT"),
        "role": os.getenv("ISSUER_ROLE", "manufacturer"),
    }


def ensure_lifecycle(passport: dict) -> dict:
    passport.setdefault("lifecycle", {"status": "draft", "events": []})
    passport["lifecycle"].setdefault("status", "draft")
    passport["lifecycle"].setdefault("events", [])
    return passport


def append_lifecycle_event(passport: dict, event: str, data: dict | None = None) -> dict:
    ensure_lifecycle(passport)
    ev = {"event": event, "timestamp": _utc_now_iso(), "data": data or {}}
    passport["lifecycle"]["events"].append(ev)
    passport["lifecycle"]["status"] = event
    return passport


# ======================================================
# PDF / IMAGE
# ======================================================

def extract_text_from_pdf(pdf_file) -> str:
    text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    return text


def image_to_base64(img: Image.Image) -> str:
    buf = BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=85)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def generate_qr_from_url(url: str) -> BytesIO:
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


# ======================================================
# GPT — ESTRAZIONE
# ======================================================

def gpt_extract_from_pdf(pdf_text: str, client, tipo: str, fields: list[str], model: str = "gpt-4o-mini") -> dict:
    """
    Output: {field: {value, confidence, explanation}}
    """
    if not pdf_text:
        return {k: {"value": "", "confidence": 0.0, "explanation": ""} for k in fields}

    template = {k: {"value": "", "confidence": 0.0, "explanation": ""} for k in fields}

    system = (
        "Sei un motore di estrazione dati rigoroso. "
        "Rispondi SOLO in JSON valido. Nessun markdown. Nessun commento. "
        "Non inventare dati. "
        "TUTTI i campi 'explanation' devono essere in italiano."
    )

    user = (
        f"Extract product passport fields for product_type={tipo}.\n"
        "Return ONLY JSON using this template:\n"
        f"{json.dumps(template, ensure_ascii=False)}\n\n"
        "PDF_TEXT:\n"
        f"{pdf_text[:20000]}"
    )

    try:
        res = client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
        )
        content = res.choices[0].message.content or "{}"
        data = json.loads(content)
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}

    out = {}
    for k in fields:
        x = data.get(k, {}) if isinstance(data.get(k, {}), dict) else {}
        out[k] = {
            "value": x.get("value", "") if x.get("value", "") is not None else "",
            "confidence": float(x.get("confidence", 0.0) or 0.0),
            "explanation": x.get("explanation", "") if x.get("explanation", "") is not None else "",
        }
    return out

def gpt_analyze_image(image_file, client: OpenAI, tipo):
    campi = ["colore","condizioni","categoria_visiva","segni_usura"]
    prompt = f"""
Analizza immagine prodotto {tipo}.
Estrai i seguenti campi: colore, condizioni, materiale_probabile, categoria_visiva, segni_usura.
Rispondi con JSON valido.
Usa null se non determinabile.
"""
    def safe_json_parse(text):
        if text.startswith("```"):
            text = "\n".join([l for l in text.splitlines() if not l.strip().startswith("```")])
        first,last = text.find("{"), text.rfind("}")
        return json.loads(text[first:last+1])
    try:
        file_id = upload_image_to_openai(image_file, client)
        resp = client.responses.create(
            model="gpt-4o",
            input=[{"role":"user","content":[{"type":"input_text","text":prompt},{"type":"input_image","file_id":file_id}]}]
        )
        data_raw = safe_json_parse(resp.output_text.strip())
        result = {}
        for c in campi:
            val = data_raw.get(c, None)
            result[c.capitalize()] = {
                "value": val if val not in [None,"","null"] else "non rilevato",
                "confidence": 0.7 if val not in [None,"","null"] else 0.0,
                "explanation": "Dato estratto da immagine" if val not in [None,"","null"] else "Non rilevabile"
            }
        return result
    except Exception:
        return {c.capitalize():{"value":"non rilevato","confidence":0.0,"explanation":"Non rilevabile"} for c in campi}

def upload_image_to_openai(image_file, client: OpenAI):
    resized = resize_image_for_vision(image_file)
    uploaded = client.files.create(file=resized, purpose="vision")
    return uploaded.id

def resize_image_for_vision(image_file, max_size=512):
    img = Image.open(image_file).convert("RGB")
    img.thumbnail((max_size, max_size))
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    buf.name = "image.jpg"
    return buf

def gpt_extract_cert_info(file_like, client, model: str = "gpt-4o-mini") -> dict:
    """
    Estrae un set base di campi certificato.
    """
    raw = file_like.read()
    if not raw:
        return {}

    is_pdf = raw[:4] == b"%PDF"
    text = extract_text_from_pdf(BytesIO(raw)) if is_pdf else ""

    schema = {
        "tipo_certificato": {"value": "", "confidence": 0.0, "explanation": ""},
        "ente_emittente": {"value": "", "confidence": 0.0, "explanation": ""},
        "numero_certificato": {"value": "", "confidence": 0.0, "explanation": ""},
        "data_emissione": {"value": "", "confidence": 0.0, "explanation": ""},
        "data_scadenza": {"value": "", "confidence": 0.0, "explanation": ""},
        "norma_riferimento": {"value": "", "confidence": 0.0, "explanation": ""},
    }

    user = (
        "Extract certificate info. Return ONLY JSON.\n"
        f"Use this schema:\n{json.dumps(schema, ensure_ascii=False)}\n\n"
        f"TEXT:\n{text[:15000]}"
    )

    try:
        res = client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[{"role": "user", "content": user}],
        )
        data = json.loads(res.choices[0].message.content or "{}")
        if not isinstance(data, dict):
            data = {}
    except Exception:
        data = {}

    out = {}
    for k, tpl in schema.items():
        v = data.get(k, tpl)
        if not isinstance(v, dict):
            v = tpl
        out[k] = {
            "value": v.get("value", ""),
            "confidence": float(v.get("confidence", 0.0) or 0.0),
            "explanation": v.get("explanation", ""),
        }
    return out


# ======================================================
# NORMALIZZAZIONE CAMPI PDF (PEF)
# ======================================================

def normalize_pdf_fields(pdf_data: dict) -> dict:
    """
    Normalizza SOLO i campi PEF, senza perdere gli altri campi PDF.
    (Niente '...': liste vere di varianti.)
    """
    FIELD_MAP = {
        "Percentuale di contenuto riciclato": [
            "Percentuale di contenuto riciclato", "% di contenuto riciclato", "% riciclato", "Contenuto riciclato"
        ],
        "Sostanze preoccupanti": [
            "Sostanze preoccupanti", "Sostanze pericolose", "SVHC", "Sostanze"
        ],
        "Energia consumata": [
            "Energia consumata", "Consumo energetico", "Energia"
        ],
        "Luogo di Produzione": [
            "Luogo di Produzione", "Luogo di produzione", "Produzione", "Made in", "Prodotto in"
        ],
        "Durabilità": [
            "Durabilità", "Durata", "Resistenza"
        ],
        "Istruzioni di riparazione": [
            "Istruzioni di riparazione", "Riparabilità", "Riparazione"
        ],
        "Parti sostituibili": [
            "Parti sostituibili", "Componenti sostituibili", "Parti di ricambio"
        ],
        "Indicazioni di smaltimento": [
            "Indicazioni di smaltimento", "Smaltimento", "Disposal"
        ],
        "Fine vita": [
            "Fine vita", "End of life", "EoL"
        ],
        "Certificazioni": [
            "Certificazioni", "Certificato", "Certificazione"
        ],
    }

    normalized = dict(pdf_data or {})

    for canonical, variants in FIELD_MAP.items():
        found = False
        for v in variants:
            if v in (pdf_data or {}) and isinstance((pdf_data or {})[v], dict) and (pdf_data or {})[v].get("value"):
                normalized[canonical] = (pdf_data or {})[v]
                found = True
                break

        if not found:
            normalized.setdefault(canonical, {"value": "", "confidence": 0.0, "explanation": ""})

    return normalized


# ======================================================
# PASSPORT CORE
# ======================================================

def initialize_passport(pid: str, tipo: str, fields: list[str]) -> dict:
    passport = {
        "id": pid,
        "product_type": tipo,
        "version": 0,
        "created_at": _utc_now_iso(),
        "last_updated_at": _utc_now_iso(),
        "sections": {"PDF": {}},
        "images": [],
        "certificates": [],
        "evidences": [],
        "physical_binding": None,
        "issuer": get_default_issuer(),
        "attestation": None,
        "digital_signature": None,
        "lifecycle": {"status": "draft", "events": []},
    }

    for f in fields:
        passport["sections"]["PDF"][f] = {"value": "", "confidence": 0.0, "explanation": ""}

    append_lifecycle_event(passport, "draft", {"reason": "initialization"})
    espr_stamp(passport, actor="manufacturer", action="initial_creation", reason="Initial creation")
    return passport


def merge_data(passport: dict, pdf_data=None, image_data=None, cert_data=None):
    if pdf_data:
        passport.setdefault("sections", {}).setdefault("PDF", {})
        passport["sections"]["PDF"].update(pdf_data)

    if image_data:
        passport.setdefault("sections", {})["Images"] = image_data

    if cert_data is not None:
        passport["certificates"] = cert_data

    append_lifecycle_event(passport, "updated", {"what": "merge_data"})
    espr_stamp(passport, actor="manufacturer", action="merge_data", reason="Merged validated data")
    return passport


def add_product_image(passport: dict, img_file, caption: str = ""):
    img = Image.open(img_file)
    passport.setdefault("images", []).append({
        "file_base64": image_to_base64(img),
        "caption": caption or ""
    })
    append_lifecycle_event(passport, "updated", {"what": "add_product_image"})
    return passport


def add_certificate_evidence(passport: dict, cert_parsed: dict, raw_bytes: bytes, filename: str = "", source: str = "uploaded_certificate"):
    evid_hash = compute_sha256_bytes(raw_bytes)
    evid_id = f"evid_{evid_hash[:16]}"

    passport.setdefault("evidences", []).append({
        "evidence_id": evid_id,
        "hash_algorithm": "SHA-256",
        "hash": evid_hash,
        "filename": filename or "",
        "source": source,
        "created_at": _utc_now_iso(),
    })

    cert_obj = cert_parsed if isinstance(cert_parsed, dict) else {"raw": {"value": str(cert_parsed)}}
    cert_obj["evidence"] = {"evidence_id": evid_id, "hash": evid_hash, "hash_algorithm": "SHA-256", "filename": filename or "", "source": source}

    passport.setdefault("certificates", []).append(cert_obj)

    append_lifecycle_event(passport, "certified", {"evidence_id": evid_id, "filename": filename})
    espr_stamp(passport, actor="manufacturer", action="add_certificate_evidence", reason=f"Added certificate evidence {evid_id}")
    return passport


def set_physical_binding(passport: dict, public_url: str, carrier: str = "qr", location: str = "product_label", tamper_risk: str = "medium"):
    passport["physical_binding"] = {
        "carrier": carrier,
        "location": location,
        "public_url": public_url,
        "generated_at": _utc_now_iso(),
        "tamper_risk": tamper_risk,
    }
    append_lifecycle_event(passport, "updated", {"physical_binding": {"carrier": carrier, "location": location}})
    espr_stamp(passport, actor="manufacturer", action="set_physical_binding", reason="Linked physical carrier to DPP")
    return passport


def merge_data_with_ecolabel(passport, pdf_file=None, image_data=None, cert_data=None, client=None):
    """
    Versione robusta e coerente col main.py:
    - se ho pdf_file+client: estraggo campi dal PDF
    - normalizzo campi PEF
    - merge con image_data/cert_data
    """
    extracted_pdf = {}
    if pdf_file and client:
        text = extract_text_from_pdf(pdf_file)
        fields = list((passport.get("sections", {}).get("PDF") or {}).keys())
        extracted_pdf = gpt_extract_from_pdf(text, client, tipo="mobile", fields=fields)

    normalized_pdf = normalize_pdf_fields(extracted_pdf)
    return merge_data(passport, pdf_data=normalized_pdf, image_data=image_data, cert_data=cert_data)


# ======================================================
# PEF + ESPR validation (stub robusto)
# ======================================================

def compute_pef_score(passport: dict) -> int:
    # Stub: puoi sostituire con il tuo scoring reale quando vuoi
    score = int(passport.get("sustainability_score") or 0)
    if score <= 0:
        score = 50
    passport["sustainability_score"] = score
    passport["sustainability_breakdown"] = passport.get("sustainability_breakdown") or {"Base": score}
    return score


def missing_pef_fields(passport: dict):
    # Stub minimal: torna lista vuota (nessun blocco)
    return []


def validate_espr_furniture(passport: dict) -> dict:
    # Stub coerente col main (missing_fields / missing_blocks / warnings / is_compliant)
    missing_fields = []
    pdf = passport.get("sections", {}).get("PDF", {}) or {}
    for k, v in pdf.items():
        if isinstance(v, dict) and str(v.get("value", "")).strip() == "" and k in PRODUCT_FIELDS.get(passport.get("product_type","mobile"), {}).get("pdf", []):
            # qui potresti applicare una true mandatory list; per ora è soft
            pass

    # blocchi minimi
    missing_blocks = []
    if not (passport.get("issuer") or {}).get("legal_name"):
        missing_blocks.append("issuer.legal_name")
    if not (passport.get("physical_binding") or {}).get("public_url"):
        missing_blocks.append("physical_binding.public_url")

    return {
        "missing_fields": missing_fields,
        "missing_blocks": missing_blocks,
        "warnings": [],
        "is_compliant": (len(missing_fields) == 0 and len(missing_blocks) == 0),
    }


def espr_stamp(passport: dict, actor: str, action: str, reason: str):
    passport.setdefault("version", 0)
    passport["version"] = int(passport["version"]) + 1
    passport["last_updated_at"] = _utc_now_iso()
    passport.setdefault("change_log", []).append({
        "version": passport["version"],
        "timestamp": passport["last_updated_at"],
        "actor": actor,
        "action": action,
        "reason": reason,
    })
    passport.setdefault("issuer", get_default_issuer())
    passport["attestation"] = {
    "type": "ESPR_PROCESS_ATTESTATION",
    "statement": (
        "Il presente Digital Product Passport è stato generato tramite la piattaforma NUVIA "
        "secondo policy ESPR dichiarate e versionate (ESPR framework – compliance enabling). "
        "La responsabilità della conformità finale del prodotto resta in capo "
        "all’operatore economico."
    ),
    "policy_version": passport.get("espr_policy_version", "v1.0"),
    "timestamp": passport["last_updated_at"],
    "responsible_actor": actor,               
}

    h = compute_passport_hash(passport)
    passport["digital_signature"] = {
        "algorithm": "SHA-256",
        "hash": h,
        "signed_at": passport["last_updated_at"],
        "signed_by": (passport.get("issuer") or {}).get("legal_name", "unknown"),
    }
    return passport


def integrate_espr_modules(passport: dict):
    """
    Versione safe: popola i campi che il Public View mostra, senza dipendenze esterne (dpp.*).
    """
    passport["jsonld"] = passport.get("jsonld") or {}
    passport["ontology"] = passport.get("ontology") or {}
    passport["eprel"] = passport.get("eprel") or {}
    passport["gs1_digital_link"] = passport.get("gs1_digital_link") or None
    passport["scip"] = passport.get("scip") or {}
    passport["espr_sections"] = passport.get("espr_sections") or {}
    passport["espr_validation"] = passport.get("espr_validation") or validate_espr_furniture(passport)
    return passport


# ======================================================
# SES SIGN (obbligatoria nel main)
# ======================================================

def sign_passport_pdf_ses_openapi(
    passport: dict,
    signer_name: str,
    signer_surname: str,
    signer_email: str,
    signer_mobile: str,
):
    """
    Versione safe: se hai OPENAPI_BEARER_PROD prova a chiamare EU-SES,
    altrimenti crea uno stub che non crasha la UI.
    """
    bearer = st.secrets.get("OPENAPI_BEARER_PROD", "")

    # Stub se manca bearer
    if not bearer:
        passport["simple_signature"] = {
            "provider": "OpenAPI",
            "type": "EU-SES",
            "status": "requested",
            "signing_urls": [],
            "raw_response": {},
        }
        return passport

    # serve pdf_document
    if "pdf_document" not in passport:
        raise RuntimeError("passport['pdf_document'] mancante — PDF non generato")

    pdf_bytes = base64.b64decode(passport["pdf_document"])
    pdf_base64 = base64.b64encode(pdf_bytes).decode("utf-8")

    url = "https://esignature.openapi.com/EU-SES"
    payload = {
        "title": "Nuvia",
        "description": "EU-SES request",
        "inputDocuments": pdf_base64,
        "signers": [{
            "name": signer_name,
            "surname": signer_surname,
            "email": signer_email,
            "mobile": signer_mobile,
            "authentication": "sms"
        }],
        "options": {"timezone": "UTC", "signatureMode": ["typed"]},
    }

    headers = {
        "Authorization": f"Bearer {bearer}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    r = requests.post(url, json=payload, headers=headers, timeout=30)
    if not r.ok:
        raise RuntimeError(f"EU-SES ERROR {r.status_code}: {r.text}")

    resp = r.json()
    data = resp.get("data", {}) if isinstance(resp, dict) else {}
    signing_urls = []
    for s in (data.get("signers") or []):
        if isinstance(s, dict) and s.get("url"):
            signing_urls.append(s["url"])

    passport["simple_signature"] = {
        "provider": "OpenAPI",
        "type": "EU-SES",
        "request_id": data.get("id"),
        "status": data.get("state"),
        "signing_urls": signing_urls,
        "raw_response": resp,
    }
    return passport


# ======================================================
# DB / STORAGE (Supabase Postgres)
# ======================================================
def _sec(name: str, default: str = "") -> str:
    try:
        if name in st.secrets:
            return str(st.secrets[name])
    except Exception:
        pass
    return str(os.getenv(name, default))


_DB_SCHEMA_READY = False

def ensure_db_schema():
    """
    Crea le 3 tabelle se non esistono.
    Se esiste già la vecchia 'passports' (solo payload/version/created/updated),
    aggiunge le nuove colonne con ALTER TABLE ... IF NOT EXISTS.
    """
    global _DB_SCHEMA_READY
    if _DB_SCHEMA_READY or not db_enabled():
        return

    ddl = """
    -- 1) passports (snapshot/listing)
    create table if not exists passports (
      passport_id text primary key,
      payload jsonb,
      version integer not null default 0,
      created_at timestamptz default now(),
      updated_at timestamptz default now()
    );

    -- evolve old passports schema safely
    alter table passports add column if not exists product_type text;
    alter table passports add column if not exists lifecycle_status text;
    alter table passports add column if not exists issuer_legal_name text;
    alter table passports add column if not exists sustainability_score integer;
    alter table passports add column if not exists pdf_present boolean default false;
    alter table passports add column if not exists cert_count integer default 0;

    -- 2) passport_fields (versioned fields)
    create table if not exists passport_fields (
      passport_id text not null,
      version integer not null,
      section text not null,
      field_name text not null,
      field_value text,
      confidence numeric,
      source text,
      updated_at timestamptz default now(),
      primary key (passport_id, version, section, field_name)
    );

    -- 3) passport_assets (images/evidences/certs/binding)
    create table if not exists passport_assets (
      asset_id text primary key,
      passport_id text not null,
      version integer not null,
      asset_type text,
      hash_sha256 text,
      filename text,
      metadata jsonb,
      created_at timestamptz default now()
    );

    -- Index utili (list + diff + lookup)
    create index if not exists idx_passports_updated_at on passports(updated_at);
    create index if not exists idx_passports_product_type on passports(product_type);
    create index if not exists idx_passports_lifecycle on passports(lifecycle_status);
    create index if not exists idx_passports_version on passports(version);

    create index if not exists idx_fields_pid_ver on passport_fields(passport_id, version);
    create index if not exists idx_fields_section on passport_fields(section);

    create index if not exists idx_assets_pid_ver on passport_assets(passport_id, version);
    create index if not exists idx_assets_type on passport_assets(asset_type);
    """

    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute(ddl)

    _DB_SCHEMA_READY = True

def _to_dict(payload):
    if payload is None:
        return None
    if isinstance(payload, dict):
        return payload
    if isinstance(payload, str):
        try:
            return json.loads(payload)
        except Exception:
            return None
    return payload


def _infer_field_source(section: str) -> str:
    s = (section or "").lower()
    if s == "pdf":
        return "pdf"
    if s == "images" or s == "image":
        return "image"
    if s == "certificates" or s == "certificate":
        return "certificate"
    return "computed"


def _safe_hash_b64_image(b64: str) -> str:
    """
    Hash stabile dei bytes reali dell'immagine.
    Se decode fallisce, hash della stringa.
    """
    try:
        raw = base64.b64decode(b64)
        return compute_sha256_bytes(raw)
    except Exception:
        return compute_sha256_bytes((b64 or "").encode("utf-8"))


def persist_passport(passport: dict, actor="manufacturer", reason="update"):
    """
    Persist completo:
      - upsert su passports (snapshot)
      - rewrite per-version su passport_fields
      - rewrite per-version su passport_assets (images, evidences, certs, binding)
    Idempotente: se richiami sulla stessa version, non duplica.
    """
    if not db_enabled():
        # fallback file (se vuoi mantenere fallback legacy, qui puoi rimetterlo)
        os.makedirs(PASSPORT_DIR, exist_ok=True)
        with open(os.path.join(PASSPORT_DIR, f"{passport['id']}.json"), "w", encoding="utf-8") as f:
            json.dump(passport, f, indent=2, ensure_ascii=False)
        return

    ensure_db_schema()

    pid = str(passport.get("id") or "").strip()
    if not pid:
        raise ValueError("passport['id'] mancante")
    ver = int(passport.get("version", 0) or 0)

    product_type = passport.get("product_type")
    lifecycle_status = (passport.get("lifecycle") or {}).get("status")
    issuer_legal_name = (passport.get("issuer") or {}).get("legal_name")
    sustainability_score = passport.get("sustainability_score")
    pdf_present = bool(passport.get("pdf_document"))
    cert_count = len(passport.get("certificates") or [])

    payload_json = json.dumps(passport, ensure_ascii=False)

    with db_conn() as conn:
        cur = conn.cursor()

        # 1) passports snapshot (1 row per passport_id)
        cur.execute("""
            insert into passports(
              passport_id, payload, version,
              product_type, lifecycle_status, issuer_legal_name,
              sustainability_score, pdf_present, cert_count,
              updated_at
            )
            values (%s, %s::jsonb, %s, %s, %s, %s, %s, %s, %s, now())
            on conflict (passport_id) do update set
              payload = excluded.payload,
              version = excluded.version,
              product_type = excluded.product_type,
              lifecycle_status = excluded.lifecycle_status,
              issuer_legal_name = excluded.issuer_legal_name,
              sustainability_score = excluded.sustainability_score,
              pdf_present = excluded.pdf_present,
              cert_count = excluded.cert_count,
              updated_at = now()
        """, (
            pid, payload_json, ver,
            product_type, lifecycle_status, issuer_legal_name,
            sustainability_score, pdf_present, cert_count
        ))

        # 2) passport_fields (rewrite this version)
        cur.execute("delete from passport_fields where passport_id=%s and version=%s", (pid, ver))

        sections = passport.get("sections") or {}
        if isinstance(sections, dict):
            for section, fields in sections.items():
                if not isinstance(fields, dict):
                    continue
                source = _infer_field_source(section)
                for fname, f in fields.items():
                    if isinstance(f, dict):
                        val = f.get("value", "")
                        conf = f.get("confidence", None)
                    else:
                        val = "" if f is None else str(f)
                        conf = None

                    cur.execute("""
                        insert into passport_fields
                        (passport_id, version, section, field_name, field_value, confidence, source, updated_at)
                        values (%s,%s,%s,%s,%s,%s,%s, now())
                    """, (
                        pid, ver,
                        str(section),
                        str(fname),
                        "" if val is None else str(val),
                        conf,
                        source
                    ))

        # 3) passport_assets (rewrite this version)
        cur.execute("delete from passport_assets where passport_id=%s and version=%s", (pid, ver))

        # 3a) images[] (base64 + caption)
        for j, img in enumerate(passport.get("images") or []):
            b64 = (img or {}).get("file_base64") or ""
            if not b64:
                continue
            img_hash = _safe_hash_b64_image(b64)
            asset_id = f"img_{pid}_{ver}_{j}_{img_hash[:12]}"
            cur.execute("""
                insert into passport_assets
                (asset_id, passport_id, version, asset_type, hash_sha256, filename, metadata, created_at)
                values (%s,%s,%s,'image',%s,%s,%s::jsonb, now())
            """, (
                asset_id, pid, ver,
                img_hash,
                (img or {}).get("filename", "") or "",
                json.dumps(img, ensure_ascii=False)
            ))

        # 3b) evidences[] (hash, filename, source...)
        for k, ev in enumerate(passport.get("evidences") or []):
            ev = ev or {}
            ev_hash = ev.get("hash") or ""
            asset_id = (ev.get("evidence_id") or "").strip()
            if not asset_id:
                asset_id = f"evid_{pid}_{ver}_{k}_{(ev_hash or 'nohash')[:12]}"
            cur.execute("""
                insert into passport_assets
                (asset_id, passport_id, version, asset_type, hash_sha256, filename, metadata, created_at)
                values (%s,%s,%s,'evidence',%s,%s,%s::jsonb, now())
            """, (
                asset_id, pid, ver,
                ev_hash,
                ev.get("filename", "") or "",
                json.dumps(ev, ensure_ascii=False)
            ))

        # 3c) certificates[] (parsed cert)
        for i, cert in enumerate(passport.get("certificates") or []):
            cert = cert or {}
            cert_hash = compute_sha256_bytes(json.dumps(cert, ensure_ascii=False).encode("utf-8"))
            asset_id = f"cert_{pid}_{ver}_{i}_{cert_hash[:12]}"
            cur.execute("""
                insert into passport_assets
                (asset_id, passport_id, version, asset_type, hash_sha256, filename, metadata, created_at)
                values (%s,%s,%s,'certificate',%s,%s,%s::jsonb, now())
            """, (
                asset_id, pid, ver,
                cert_hash,
                cert.get("filename", "") if isinstance(cert, dict) else "",
                json.dumps(cert, ensure_ascii=False)
            ))

        # 3d) physical_binding (qr/link)
        if passport.get("physical_binding"):
            pb = passport["physical_binding"]
            pb_hash = compute_sha256_bytes(json.dumps(pb, ensure_ascii=False).encode("utf-8"))
            asset_id = f"bind_{pid}_{ver}_{pb_hash[:12]}"
            cur.execute("""
                insert into passport_assets
                (asset_id, passport_id, version, asset_type, hash_sha256, filename, metadata, created_at)
                values (%s,%s,%s,'binding',%s,%s,%s::jsonb, now())
            """, (
                asset_id, pid, ver,
                pb_hash,
                "",
                json.dumps(pb, ensure_ascii=False)
            ))


def load_passport(pid: str):
    """
    DB-first, fallback file.
    Nota: qui ritorni ancora il payload completo (compatibilità con UI attuale).
    """
    if not db_enabled():
        path = os.path.join(PASSPORT_DIR, f"{pid}.json")
        if not os.path.exists(path):
            return None
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    ensure_db_schema()

    with db_conn() as conn:
        cur = conn.cursor()
        cur.execute("select payload from passports where passport_id=%s", (pid,))
        row = cur.fetchone()
        return _to_dict(row[0]) if row else None


def db_list_passports_latest(
    search: str = "",
    product_type: str = "ALL",
    lifecycle: str = "ALL",
    min_version: int = 1,
    has_pdf: bool = False,
    has_cert: bool = False,
    sort_col: str = "updated_at",
    sort_asc: bool = False,
    limit: int = 5000,
) -> pd.DataFrame:
    """
    LIST VIEW basata SOLO su passports (no JSON traversal).
    Colonne:
      id | product_type | version | lifecycle | pdf_present | cert_count | created_at | updated_at
    """
    ensure_db_schema()

    allowed_sort = {"created_at", "updated_at", "version"}
    if sort_col not in allowed_sort:
        sort_col = "updated_at"
    order = "ASC" if sort_asc else "DESC"

    where = ["1=1"]
    params = []

    if search:
        where.append("""
            (
              passport_id ILIKE %s
              OR coalesce(issuer_legal_name,'') ILIKE %s
              OR coalesce(product_type,'') ILIKE %s
            )
        """)
        like = f"%{search}%"
        params += [like, like, like]

    if product_type and product_type != "ALL":
        where.append("product_type = %s")
        params.append(product_type)

    if lifecycle and lifecycle != "ALL":
        where.append("lifecycle_status = %s")
        params.append(lifecycle)

    if min_version and int(min_version) > 1:
        where.append("version >= %s")
        params.append(int(min_version))

    if has_pdf:
        where.append("pdf_present = true")

    if has_cert:
        where.append("cert_count > 0")

    sql = f"""
        select
            passport_id as id,
            product_type,
            version,
            lifecycle_status as lifecycle,
            pdf_present,
            cert_count,
            created_at,
            updated_at
        from passports
        where {" and ".join(where)}
        order by {sort_col} {order}
        limit %s
    """
    params.append(int(limit))

    with db_conn() as conn:
        return pd.read_sql(sql, conn, params=params)
        

def compute_diff_fields(
    passport_id: str,
    v_old: int,
    v_new: int
) -> pd.DataFrame:
    ensure_db_schema()

    sql = """
    select
        f_old.section,
        f_old.field_name,
        f_old.field_value as old_value,
        f_new.field_value as new_value
    from passport_fields f_old
    join passport_fields f_new
      on f_old.passport_id = f_new.passport_id
     and f_old.section = f_new.section
     and f_old.field_name = f_new.field_name
    where
        f_old.passport_id = %s
        and f_old.version = %s
        and f_new.version = %s
        and coalesce(f_old.field_value,'') <> coalesce(f_new.field_value,'')
    order by f_old.section, f_old.field_name
    """

    with db_conn() as conn:
        return pd.read_sql(sql, conn, params=[passport_id, v_old, v_new])

def db_enabled() -> bool:
    return bool(_sec("SUPABASE_DB_URL"))


@contextmanager
def db_conn():
    url = _sec("SUPABASE_DB_URL")
    if not url:
        raise RuntimeError("SUPABASE_DB_URL mancante")

    # prova psycopg2
    try:
        import psycopg2
        conn = psycopg2.connect(url)
        conn.autocommit = False
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return
    except ImportError:
        pass

    # fallback pg8000
    import pg8000
    from urllib.parse import urlparse

    u = urlparse(url)
    conn = pg8000.connect(
        user=u.username,
        password=u.password,
        host=u.hostname,
        port=u.port or 5432,
        database=(u.path or "/postgres").lstrip("/"),
    )

    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def save_passport_to_excel_append(passport: dict):
    """
    Salva/append su Excel legacy per Archivio + Diff versioni.
    Sheet:
      - passport  (meta per versione)
      - fields    (campi per versione)  ✅ include 'version'
      - images    (immagini per versione) ✅ include 'version'
    """
    os.makedirs(PASSPORT_DIR, exist_ok=True)

    pid = str(passport.get("id") or "").strip()
    ver = int(passport.get("version") or 0)

    issuer = passport.get("issuer") or {}
    lifecycle = passport.get("lifecycle") or {}
    pb = passport.get("physical_binding") or {}

    # ---------------------------
    # 1) SHEET: passport (meta)
    # ---------------------------
    meta_row = {
        "id": pid,
        "version": ver,
        "product_type": passport.get("product_type"),
        "created_at": passport.get("created_at"),
        "last_updated_at": passport.get("last_updated_at"),
        "issuer_legal_name": issuer.get("legal_name"),
        "lifecycle": lifecycle.get("status"),
        "pdf_present": bool(passport.get("pdf_document")),
        "cert_count": len(passport.get("certificates") or []),
        "binding_public_url": pb.get("public_url"),
    }
    df_passport = pd.DataFrame([meta_row])

    # ---------------------------
    # 2) SHEET: fields  ✅ version qui
    # ---------------------------
    fields_rows = []
    sections = passport.get("sections") or {}
    if isinstance(sections, dict):
        for section, fields in sections.items():
            if not isinstance(fields, dict):
                continue
            for field_name, f in fields.items():
                if isinstance(f, dict):
                    val = f.get("value")
                else:
                    val = f
                fields_rows.append({
                    "passport_id": pid,
                    "version": ver,               # ✅ FONDAMENTALE PER IL DIFF
                    "section": section,
                    "field_name": field_name,
                    "value": "" if val is None else str(val),
                })
    df_fields = pd.DataFrame(fields_rows)

    # ---------------------------
    # 3) SHEET: images ✅ version anche qui
    # ---------------------------
    images_rows = []
    for img in (passport.get("images") or []):
        images_rows.append({
            "passport_id": pid,
            "version": ver,                   # ✅ utile anche per diff media
            "file_base64": img.get("file_base64"),
            "caption": img.get("caption", ""),
        })
    df_images = pd.DataFrame(images_rows)

    # ---------------------------
    # 4) CREA FILE SE NON ESISTE
    # ---------------------------
    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            df_passport.to_excel(writer, sheet_name="passport", index=False)
            df_fields.to_excel(writer, sheet_name="fields", index=False)
            df_images.to_excel(writer, sheet_name="images", index=False)
        return

    # ---------------------------
    # 5) APPEND SE ESISTE
    # ---------------------------
    book = load_workbook(EXCEL_FILE)

    def _append_df(df: pd.DataFrame, sheet: str):
        if df is None or df.empty:
            return

        # se sheet non esiste, lo creiamo scrivendo da riga 0 con header
        if sheet not in book.sheetnames:
            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
            return

        startrow = book[sheet].max_row
        # se max_row==1 ma il foglio è vuoto con solo header, ok: append da 1
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=startrow)

    _append_df(df_passport, "passport")
    _append_df(df_fields, "fields")
    _append_df(df_images, "images")



def generate_pdf_from_html(html: str) -> bytes:
    API_KEY = st.secrets["OPENAPI_PDF_TOKEN"]
    endpoint = "https://pdf.openapi.it/base"

    payload = {
        "html": html,
        "format": "A4",
        "margin": "20px",
        "printBackground": True,
        "pageRanges": "1-"
    }

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }

    resp = requests.post(endpoint, json=payload, headers=headers)
    resp.raise_for_status()
    return resp.content

def load_image_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

def generate_passport_html(passport: dict, qr_base64: str = None) -> str:
    import json
    from datetime import datetime

    # ---------------------------------------------------------
    # ASSETS
    # ---------------------------------------------------------
    logo_base64 = load_image_base64("functions/logo_nuvia.jpeg")
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    status = (passport.get("lifecycle") or {}).get("status", "")
    sustainability_score = passport.get("sustainability_score", 0)

    pdf = passport.get("sections", {}).get("PDF", {})
    eprel = passport.get("eprel") or {}
    gs1 = passport.get("gs1_digital_link") or "—"
    scip = passport.get("scip") or {"scip:substances": []}
    scip_substances = scip.get("scip:substances", [])
    certificates = passport.get("certificates", [])

    issuer = passport.get("issuer") or {}
    attestation = passport.get("attestation") or {}
    signature = passport.get("digital_signature") or {}

    # ---------------------------------------------------------
    # SEMAFORO + GAUGE
    # ---------------------------------------------------------
    def traffic_light_from_score(score: int):
        try:
            s = int(score)
        except:
            s = 0

        if s >= 71:
            return "#27CC6C", "Low impact"
        elif s >= 41:
            return "#F6C23C", "Medium impact"
        else:
            return "#E63946", "High impact"

    color, label = traffic_light_from_score(sustainability_score)
    angle = int((sustainability_score / 100) * 360)

    def gauge_html():
        return f"""
        <div class="gauge-wrapper">
            <div class="gauge" style="
                background: conic-gradient({color} {angle}deg, #E0E0E0 {angle}deg);
            ">
                <div class="gauge-inner">
                    <div class="gauge-value">{sustainability_score}%</div>
                </div>
            </div>
            <div class="gauge-label" style="color:{color};">{label}</div>
        </div>
        """

    # ---------------------------------------------------------
    # COVER PAGE (logo ridotto)
    # ---------------------------------------------------------
    cover_html = f"""
    <div class="cover">
        <img src="data:image/jpeg;base64,{logo_base64}" class="cover-logo"/>
        <h1 class="cover-title">📘 Digital Product Passport</h1>
        <h2 class="cover-product">{passport.get("product_name") or passport.get("id","")}</h2>

        <div class="cover-meta">
            <div><b>ID:</b> {passport.get("id","")}</div>
            <div><b>Version:</b> {passport.get("version","")}</div>
            <div><b>Status:</b> {status}</div>
        </div>

        {f"<div class='cover-qr-frame'><img class='cover-qr' src='data:image/png;base64,{qr_base64}' /></div>" if qr_base64 else ""}

        <div class="cover-footer">
            Generated by Nuvia DPP System<br/>{generated_at}
        </div>
    </div>
    <div class="page-break"></div>
    """

    # ---------------------------------------------------------
    # SUMMARY SECTION
    # ---------------------------------------------------------
    summary_html = f"""
    <div class="section summary">
        <h2>📄 Passport Summary</h2>
        <table class="data-table">
            <tr><td class='field-name'>Product Name</td><td>{passport.get("product_name")}</td></tr>
            <tr><td class='field-name'>Product ID</td><td>{passport.get("id")}</td></tr>
            <tr><td class='field-name'>Environmental Impact</td><td>{gauge_html()}</td></tr>
            <tr><td class='field-name'>Status</td><td>{status}</td></tr>
        </table>
    </div>
    """

    # ---------------------------------------------------------
    # BREAKDOWN PEF
    # ---------------------------------------------------------
    pef_breakdown = passport.get("sustainability_breakdown", {})
    pef_rows = "".join(
        f"<tr><td class='field-name'>{k}</td><td>{v}</td></tr>"
        for k, v in pef_breakdown.items()
    )

    breakdown_html = f"""
    <div class="section">
        <h2>📊 Environmental Impact Breakdown</h2>
        <table class="data-table">
            {pef_rows}
        </table>
    </div>
    """

    # ---------------------------------------------------------
    # SEZIONI ESPR — con icone
    # ---------------------------------------------------------
    espr_html = f"""
    <div class="section">
        <h2>🔖 1. Product Identity</h2>
        <table class="data-table">
            <tr><td class='field-name'>Product Name</td><td>{pdf.get("Nome prodotto",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Model Number</td><td>{pdf.get("Numero di modello",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Manufacturer</td><td>{pdf.get("Produttore",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Place of Production</td><td>{pdf.get("Luogo di Produzione",{}).get("value","—")}</td></tr>
        </table>
    </div>

    <div class="section">
        <h2>🧱 2. Materials & Substances</h2>
        <table class="data-table">
            <tr><td class='field-name'>Materials</td><td>{pdf.get("Materiali/componenti utilizzati",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Recycled Content</td><td>{pdf.get("Percentuale di contenuto riciclato",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Hazardous Substances (SCIP/ECHA)</td>
                <td>
    """

    if scip_substances:
        for s in scip_substances:
            name = s.get("name","—")
            uri = s.get("echa_uri")
            espr_html += f"• {name}<br>"
            if uri:
                espr_html += f"<small>{uri}</small><br>"
    else:
        espr_html += "None declared"

    espr_html += """
                </td>
            </tr>
        </table>
    </div>

    <div class="section">
        <h2>🛠️ 3. Repairability & Durability</h2>
        <table class="data-table">
            <tr><td class='field-name'>Durability</td><td>{}</td></tr>
            <tr><td class='field-name'>Repair Instructions</td><td>{}</td></tr>
            <tr><td class='field-name'>Replaceable Parts</td><td>{}</td></tr>
        </table>
    </div>
    """.format(
        pdf.get("Durabilità",{}).get("value","—"),
        pdf.get("Istruzioni di riparazione",{}).get("value","—"),
        pdf.get("Parti sostituibili",{}).get("value","—")
    )

    espr_html += f"""
    <div class="section">
        <h2>♻️ 4. End of Life</h2>
        <table class="data-table">
            <tr><td class='field-name'>Disposal Instructions</td><td>{pdf.get("Indicazioni di smaltimento",{}).get("value","—")}</td></tr>
        </table>
    </div>

    <div class="section">
        <h2>📜 5. Certifications</h2>
    """

    for c in certificates:
        espr_html += f"""
        <table class="data-table">
            <tr><td class='field-name'>Certificate</td><td>{c.get("tipo_certificato",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Issuer</td><td>{c.get("ente_emittente",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Number</td><td>{c.get("numero_certificato",{}).get("value","—")}</td></tr>
            <tr><td class='field-name'>Standard</td><td>{c.get("norma_riferimento",{}).get("value","—")}</td></tr>
        </table>
        """

    espr_html += f"""
    <div class="section">
        <h2>💡 6. EPREL</h2>
        <table class="data-table">
            <tr><td class='field-name'>Energy Class</td><td>{eprel.get("eprel:energyClass","—")}</td></tr>
        </table>
    </div>

    <div class="section">
        <h2>🌐 7. GS1 Digital Link</h2>
        <table class="data-table">
            <tr><td class='field-name'>GS1 URI</td><td>{gs1}</td></tr>
        </table>
    </div>

    <div class="section">
        <h2>🔗 8. QR Code</h2>
        <div class="qr-inline">
            <img src="data:image/png;base64,{qr_base64}" width="160"/>
        </div>
    </div>

    <div class="section">
        <h2>✒️ 9. Signature & Integrity</h2>
        <table class="data-table">
            <tr><td class='field-name'>Issuer</td><td>{issuer.get("legal_name","—")}</td></tr>
            <tr><td class='field-name'>Attestation</td><td>{attestation.get("statement","—")}</td></tr>
            <tr><td class='field-name'>Document Hash</td><td>{signature.get("hash","—")}</td></tr>
        </table>
    </div>
    """

    # ---------------------------------------------------------
    # LIFECYCLE EVENTS
    # ---------------------------------------------------------
    lifecycle_html = ""
    events = (passport.get("lifecycle") or {}).get("events") or []

    if events:
        rows = ""
        for ev in events:
            rows += f"""
            <tr>
                <td>{ev.get('event')}</td>
                <td>{ev.get('timestamp')}</td>
                <td><pre>{json.dumps(ev.get('data', {}), ensure_ascii=False, indent=2)}</pre></td>
            </tr>
            """

        lifecycle_html = f"""
        <div class="page-break"></div>
        <div class="section">
            <h2>📅 Lifecycle Events</h2>
            <table class="data-table">
                <tr><th>Event</th><th>Timestamp</th><th>Details</th></tr>
                {rows}
            </table>
        </div>
        """

    # ---------------------------------------------------------
    # CHANGE LOG
    # ---------------------------------------------------------
    changelog_html = ""
    if passport.get("change_log"):
        rows = ""
        for log in passport["change_log"]:
            rows += f"""
            <tr>
                <td>{log.get('version')}</td>
                <td>{log.get('timestamp')}</td>
                <td>{log.get('actor')}</td>
                <td>{log.get('action')}</td>
                <td>{log.get('reason')}</td>
            </tr>
            """

        changelog_html = f"""
        <div class="section">
            <h2>📝 Change Log</h2>
            <table class="data-table">
                <tr><th>Version</th><th>Date</th><th>Actor</th><th>Action</th><th>Reason</th></tr>
                {rows}
            </table>
        </div>
        """

    # ---------------------------------------------------------
    # IMAGES
    # ---------------------------------------------------------
    images_html = ""
    images = passport.get("images", [])

    if images:
        cards = ""
        for img in images:
            b64 = img.get("file_base64")
            if not b64:
                continue

            cards += f"""
            <div class="image-card">
                <img src="data:image/jpeg;base64,{b64}" />
                <div class="caption">{img.get("caption","")}</div>
            </div>
            """

        images_html = f"""
        <div class="page-break"></div>
        <div class="section">
            <h2>🖼️ Product Visual Documentation</h2>
            <div class="image-grid">{cards}</div>
        </div>
        """

    # ---------------------------------------------------------
    # ABOUT + LEGAL (logo ridotto)
    # ---------------------------------------------------------
    about_html = f"""
    <div class="page-break"></div>
    <div class="about">
        <img src="data:image/jpeg;base64,{logo_base64}" class="about-logo"/>
        <h1>About Nuvia</h1>
        <p>Nuvia è una piattaforma digitale per la generazione automatizzata del Digital Product Passport, progettata per supportare la tracciabilità e la sostenibilità dei prodotti in conformità alle normative europee.</p>
        <p>Supportiamo il Digital Product Passport secondo regolamento ESPR 2024/1781.</p>
        <h2>Mission</h2>
        <p>Abilitare trasparenza e sostenibilità lungo l’intero ciclo di vita dei prodotti.</p>
        <h2>Contatti</h2>
        <p>Email: informazioni.nuvia@gmail.com</p>
        <p>Web: https://nuviadpp.com</p>
    </div>
    """

    legal_html = """
    <div class="page-break"></div>
    <div class="legal">
        <h1>Legal Notice</h1>
        <p>Questo documento è generato automaticamente dal sistema Nuvia DPP.</p>
        <p>I dati sono forniti dal produttore sotto la propria responsabilità.</p>
        <p>Nuvia non è responsabile per errori o omissioni.</p>
        <p>© Nuvia S.r.l. – Tutti i diritti riservati</p>
    </div>
    """

    # ---------------------------------------------------------
    # FINAL HTML (solo aggiunta classi logo)
    # ---------------------------------------------------------
    html = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            @page {{ margin: 40px; @bottom-right {{ content: "Page " counter(page); font-size: 9px; }} }}
            body {{ font-family: "Inter", Arial, sans-serif; color: #3D0F06; background: #F6F8F8; }}
            .page-break {{ page-break-before: always; }}
            .section {{ margin-bottom:25px; background:#F6F8F8; padding:15px; border-radius:6px; }}
            h2 {{ border-bottom:2px solid #23CE6B; padding-bottom:5px; color:#36120D; }}
            table {{ width:100%; border-collapse: collapse; font-size:12px; }}
            th, td {{ border:1px solid #27CC6C; padding:6px; }}
            .field-name {{ font-weight:bold; background:#F6F8F8; width:30%; color:#3D0F06; }}
            .image-grid {{ display:flex; flex-wrap:wrap; gap:10px; }}
            .image-card {{ width:48%; border:1px solid #27CC6C; }}
            .image-card img {{ width:100%; }}
            .caption {{ font-size:10px; text-align:center; }}

            /* LOGHI RIDOTTI */
            .cover-logo {{
                width: 190px;
                margin-bottom: 85px;
            }}

            .about-logo {{
                width: 120px;
                margin-bottom: 20px;
            }}
        </style>
    </head>
    <body>
        {cover_html}
        {summary_html}
        {breakdown_html}
        {espr_html}
        {lifecycle_html}
        {changelog_html}
        {images_html}
        {about_html}
        {legal_html}
    </body>
    </html>
    """

    return html
    




